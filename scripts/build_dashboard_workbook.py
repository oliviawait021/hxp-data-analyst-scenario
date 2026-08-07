"""
Assemble the final multi-tab workbook: raw data, both transformations, and
a Dashboard sheet, all in one file ready to open directly in Google Sheets.

Sheets (in tab order):
  1. Dashboard              - KPI cards + charts + sample warm-lead quotes
  2. Chart Data             - small COUNTIF-driven tables the charts read from
  3. Raw Data               - untouched copy of the original survey export
  4. Cleaned + Categorized  - output of transform_data.py
  5. Insights               - output of derive_insights.py
  6. Insights Summary       - output of derive_insights.py (headline metrics)

KPI cards and Chart Data counts are Excel/Sheets formulas (COUNTIF/COUNTIFS
against the Insights sheet), not hardcoded numbers - if a category gets
hand-edited later in Sheets, the dashboard updates itself.

Usage: python3 scripts/build_dashboard_workbook.py
"""

import csv
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = ROOT / "Data Team Scenario.xlsx"
CLEANED_CSV = ROOT / "data" / "processed" / "parent_survey_cleaned.csv"
INSIGHTS_CSV = ROOT / "data" / "processed" / "parent_survey_insights.csv"
SUMMARY_CSV = ROOT / "data" / "processed" / "insights_summary.csv"
OUTPUT_XLSX = ROOT / "dashboard" / "HXP Parent Builder Dashboard.xlsx"

# Palette (validated categorical set - see dataviz skill references/palette.md)
BLUE = "2A78D6"
ORANGE = "EB6834"
MUTED = "898781"
LIGHT_MUTED = "C3C2B7"
INK = "0B0B0B"
SECONDARY_INK = "52514E"
CARD_FILL = "F9F9F7"
BORDER_COLOR = "C3C2B7"
HEADER_FILL = "2A78D6"

THIN_BORDER = Border(*(Side(style="thin", color=BORDER_COLOR) for _ in range(4)))

# Fixed label order for each Chart Data table (labels are static; counts are
# live COUNTIF formulas, so this only controls display order).
NOT_GOING_CATEGORY_ORDER = [
    "Cost / Finances", "Work Schedule / Time Off", "Wants Child to Have Independent Experience",
    "Other Kids / Family at Home", "Child Didn't Want Parent to Come",
    "Another Family Member Going Instead", "Other / Uncategorized",
    "Not a Traveler / Personal Preference", "Not Aware It Was an Option",
    "Other Life Commitment", "Health / Physical / Age", "Family's Own Choice (No Elaboration)",
    "Trip / Spot Was Already Full", "Not Needed",
]
GOING_CATEGORY_ORDER = [
    "Wanted Shared Experience with Child", "Previous Positive Experience",
    "Faith / Service Motivation", "Child/Spouse Asked Them to Go",
    "Family Member Was the Reason (No Elaboration)", "Other / Uncategorized",
    "Love of Travel / Culture", "Trust in the HXP Program",
    "Concerned About Child Going Alone / Safety",
    "Securing/Ensuring Child's Spot (Registration Timing)",
    "Personally Invited by HXP Staff/Trip Leader",
]
CONTROLLABILITY_ORDER = [
    "HXP Can Directly Fix", "HXP Can Influence", "Family's Own Choice", "Unknown",
]
# The subset of NOT_GOING categories mapped to "HXP Can Directly Fix" or
# "HXP Can Influence" in derive_insights.py's CONTROLLABILITY dict, ranked.
CONTROLLABLE_REASON_ORDER = [
    "Cost / Finances", "Work Schedule / Time Off", "Not a Traveler / Personal Preference",
    "Not Aware It Was an Option", "Health / Physical / Age", "Trip / Spot Was Already Full",
]
# "Easy conversions": controllable reasons that need zero persuasion to fix -
# purely information or operations, not a budget or schedule change. A
# judgment call layered on top of controllability - revisit if the Parent
# Builder team reads it differently (e.g. whether "Not a Traveler" belongs
# here too is arguable; left out since it's an attitude shift, not a fact fix).
EASY_CONVERSION_ORDER = [
    "Not Aware It Was an Option", "Trip / Spot Was Already Full",
]
WOULD_HELP_ORDER = [
    "Cost / Financial Assistance", "Time / Scheduling", "Childcare for Other Kids",
    "Waiting for Right Timing (kids' ages)", "Other / Uncategorized",
    "No Specific Ask / Undecided", "More Information / Awareness", "Health",
    "Registration / Spot Availability",
]


def style_header_row(ws, row=1, ncols=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_csv_sheet(wb, sheet_name, csv_path, col_widths=None, type_converters=None):
    """type_converters: {0-based column index: fn(str) -> typed value}, so
    e.g. booleans/numbers land as real Excel types instead of text - matters
    for any cell a COUNTIF/COUNTIFS formula elsewhere needs to match on type.
    """
    ws = wb.create_sheet(sheet_name)
    type_converters = type_converters or {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        ws.append(next(reader))  # header, no conversion
        for row in reader:
            converted = [type_converters[i](v) if i in type_converters else v
                         for i, v in enumerate(row)]
            ws.append(converted)
    ncols = ws.max_column
    style_header_row(ws, ncols=ncols)
    ws.auto_filter.ref = ws.dimensions
    widths = col_widths or [22] * ncols
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def write_raw_data_sheet(wb):
    ws = wb.create_sheet("Raw Data")
    src = openpyxl.load_workbook(SOURCE_XLSX, data_only=True).active
    for row in src.iter_rows(values_only=True):
        ws.append(row)
    ncols = ws.max_column
    style_header_row(ws, ncols=ncols)
    ws.auto_filter.ref = ws.dimensions
    for i in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 45
    return ws


def add_kpi_card(ws, top_left_col, row, width, label, formula, number_format, color=INK):
    col_letters = [get_column_letter(top_left_col + i) for i in range(width)]
    value_range = f"{col_letters[0]}{row}:{col_letters[-1]}{row + 2}"
    label_range = f"{col_letters[0]}{row + 3}:{col_letters[-1]}{row + 3}"

    ws.merge_cells(value_range)
    vcell = ws[f"{col_letters[0]}{row}"]
    vcell.value = formula
    vcell.number_format = number_format
    vcell.font = Font(size=26, bold=True, color=color)
    vcell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(label_range)
    lcell = ws[f"{col_letters[0]}{row + 3}"]
    lcell.value = label
    lcell.font = Font(size=10, color=SECONDARY_INK)
    lcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(row, row + 4):
        for c in range(top_left_col, top_left_col + width):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CARD_FILL)
            ws.cell(row=r, column=c).border = THIN_BORDER


def build_chart_data_sheet(wb):
    ws = wb.create_sheet("Chart Data")
    ws.sheet_state = "visible"
    ws["A1"] = "Backing data for the Dashboard charts - safe to leave visible, not meant for managers to read directly."
    ws["A1"].font = Font(italic=True, color=SECONDARY_INK, size=9)

    def table(title, start_row, col_a_header, categories, count_col_letter, formula_fn):
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=11)
        ws.cell(row=start_row + 1, column=1, value=col_a_header).font = Font(bold=True)
        ws.cell(row=start_row + 1, column=2, value="Count").font = Font(bold=True)
        for i, cat in enumerate(categories):
            r = start_row + 2 + i
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=formula_fn(cat))
        return start_row + 1, start_row + 1 + len(categories)  # header_row, last_data_row

    ngo_header, ngo_last = table(
        "Why Parents Don't Go", 3, "Category", NOT_GOING_CATEGORY_ORDER, "B",
        lambda cat: f'=COUNTIF(Insights!$F$2:$F$1190,"{cat}")',
    )
    go_header, go_last = table(
        "Why Parents Do Go", ngo_last + 3, "Category", GOING_CATEGORY_ORDER, "B",
        lambda cat: f'=COUNTIF(Insights!$C$2:$C$1190,"{cat}")',
    )
    ctrl_header, ctrl_last = table(
        "Controllability", go_last + 3, "Bucket", CONTROLLABILITY_ORDER, "B",
        lambda cat: f'=COUNTIF(Insights!$I$2:$I$1190,"{cat}")',
    )
    wh_header, wh_last = table(
        "Would-Help Themes (Maybe segment)", ctrl_last + 3, "Theme", WOULD_HELP_ORDER, "B",
        lambda cat: f'=COUNTIF(Insights!$L$2:$L$1190,"{cat}")',
    )
    controllable_header, controllable_last = table(
        "Reasons Parents Don't Go That HXP Can Control", wh_last + 3, "Category",
        CONTROLLABLE_REASON_ORDER, "B",
        lambda cat: f'=COUNTIF(Insights!$F$2:$F$1190,"{cat}")',
    )
    easy_header, easy_last = table(
        "Easy Conversions (low-effort fixes within the controllable set)", controllable_last + 3,
        "Category", EASY_CONVERSION_ORDER, "B",
        lambda cat: f'=COUNTIF(Insights!$F$2:$F$1190,"{cat}")',
    )

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    return ws, {
        "not_going": (ngo_header, ngo_last),
        "going": (go_header, go_last),
        "controllability": (ctrl_header, ctrl_last),
        "would_help": (wh_header, wh_last),
        "controllable_reasons": (controllable_header, controllable_last),
        "easy_conversions": (easy_header, easy_last),
    }


def make_bar_chart(ws_data, header_row, last_row, title, single_color=None, point_colors=None):
    chart = BarChart()
    chart.type = "bar"  # horizontal bars - reads better with long category labels
    chart.title = title
    chart.style = 10
    chart.legend = None
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.height = 8
    chart.width = 18

    data = Reference(ws_data, min_col=2, min_row=header_row, max_row=last_row)
    cats = Reference(ws_data, min_col=1, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    series = chart.series[0]
    series.graphicalProperties.solidFill = single_color or BLUE
    if point_colors:
        series.data_points = [DataPoint(idx=i) for i in range(len(point_colors))]
        for i, color in enumerate(point_colors):
            series.data_points[i].graphicalProperties.solidFill = color

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    return chart


def get_soft_no_quotes(n=5):
    with open(INSIGHTS_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    quotes = [r["Reason Not To Go"] for r in rows if r["Not Going - Soft No Signal"] == "True"]
    picked = []
    for q in quotes:
        if len(picked) >= n:
            break
        picked.append(q if len(q) <= 150 else q[:147] + "...")
    return picked


def build_dashboard_sheet(wb, chart_ranges):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    ws["B2"] = "HXP Parent Builder Survey — Dashboard"
    ws["B2"].font = Font(size=18, bold=True, color=INK)
    ws["B3"] = "1,189 parent responses collected before this year's trips"
    ws["B3"].font = Font(size=11, color=SECONDARY_INK, italic=True)

    ctrl_header, ctrl_last = chart_ranges["controllability"]
    ctrl_data_range = f"'Chart Data'!B{ctrl_header + 1}:B{ctrl_last}"

    add_kpi_card(ws, 2, 5, 3, "Total Responses", "=COUNTA(Insights!A2:A1190)", "#,##0")
    add_kpi_card(ws, 5, 5, 3, "Not Going (No)",
                 '=COUNTIF(Insights!A2:A1190,"No")/COUNTA(Insights!A2:A1190)', "0.0%", color=ORANGE)
    # Denominator reuses the already-verified Chart Data controllability counts
    # (sum of all 4 buckets = every "No" row that has a stated reason) rather
    # than a COUNTIFS(...,"<>") not-blank idiom, which the formula verifier
    # used to check this workbook could not reliably confirm. Range is derived
    # from chart_ranges, not hardcoded - the controllability table's row
    # position shifts whenever another table above it changes length.
    add_kpi_card(ws, 8, 5, 3, "Of Those 'No's, % HXP Can Control",
                 "=(COUNTIF(Insights!I2:I1190,\"HXP Can Directly Fix\")+COUNTIF(Insights!I2:I1190,\"HXP Can Influence\"))"
                 f"/SUM({ctrl_data_range})",
                 "0.0%", color=BLUE)
    add_kpi_card(ws, 11, 5, 3, "'No's That Are Family Preference",
                 '=COUNTIF(Insights!I2:I1190,"Family\'s Own Choice")', "#,##0", color=MUTED)
    add_kpi_card(ws, 14, 5, 3, "Warm Leads (Soft 'No' Signal)",
                 "=COUNTIF(Insights!J2:J1190,TRUE)", "#,##0", color=ORANGE)

    ws_data = wb["Chart Data"]
    go_header, go_last = chart_ranges["going"]
    wh_header, wh_last = chart_ranges["would_help"]
    controllable_header, controllable_last = chart_ranges["controllable_reasons"]
    easy_header, easy_last = chart_ranges["easy_conversions"]

    chart1 = make_bar_chart(ws_data, controllable_header, controllable_last,
                             "Reasons Parents Don't Go That HXP Can Control", single_color=BLUE)
    ws.add_chart(chart1, "B11")

    chart2 = make_bar_chart(ws_data, easy_header, easy_last,
                             "Easy Conversions (low-effort fixes)", single_color=ORANGE)
    ws.add_chart(chart2, "L11")

    chart3 = make_bar_chart(ws_data, go_header, go_last, "Why Parents Do Go", single_color=BLUE)
    ws.add_chart(chart3, "B29")

    chart4 = make_bar_chart(ws_data, wh_header, wh_last, "What Would Help \"Maybe\" Parents Decide",
                             single_color=ORANGE)
    ws.add_chart(chart4, "L29")

    quote_row = 47
    ws.cell(row=quote_row, column=2, value="Sample \"Warm Lead\" Quotes (parents who said no, but hinted they'd reconsider)")
    ws.cell(row=quote_row, column=2).font = Font(bold=True, size=11)
    for i, quote in enumerate(get_soft_no_quotes()):
        cell = ws.cell(row=quote_row + 1 + i, column=2, value=f"“{quote}”")
        cell.font = Font(italic=True, color=SECONDARY_INK)
        ws.merge_cells(start_row=quote_row + 1 + i, start_column=2, end_row=quote_row + 1 + i, end_column=9)

    for col, w in zip("BCDEFGHIJKLMNOP", [11] * 16):
        ws.column_dimensions[col].width = w
    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    ws_data, chart_ranges = build_chart_data_sheet(wb)
    write_raw_data_sheet(wb)
    write_csv_sheet(wb, "Cleaned + Categorized", CLEANED_CSV,
                     col_widths=[10, 40, 30, 35, 40, 30, 35, 40])
    # Column 9 = "Not Going - Soft No Signal" (bool), column 10 = "Not Going -
    # Reason Stack Count" (int) - convert from CSV text so the KPI/COUNTIF
    # formulas that reference them can match on real Excel types.
    write_csv_sheet(wb, "Insights", INSIGHTS_CSV,
                     col_widths=[10, 40, 30, 35, 40, 30, 35, 40, 24, 14, 12, 24, 30],
                     type_converters={9: lambda v: v == "True", 10: lambda v: int(v) if v else 0})
    write_csv_sheet(wb, "Insights Summary", SUMMARY_CSV, col_widths=[24, 45, 12])

    build_dashboard_sheet(wb, chart_ranges)

    wb.active = 0  # Dashboard tab shows first when the file opens
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
